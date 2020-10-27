from openpyxl import Workbook, load_workbook

from . import *


class CsvGenerator:
    """
    Generate CSV in File Storage directory
    """

    def __init__(self, data, task):
        """
        Setup for csv
        The params should be handled by the Task class automatically
        However here is the basic Dictionary Needs
        { "header": [List of items for the name of each column],
          "file_name": "name_of_file.csv", - The File name for the csv
          "file_path": "path\\to\\file, - The File Path to store the file in
          "dynamic_name": "%y-%m-%d", - This is a date format in python datetime formatting
          "after_before": "after" - If the dynamic name should show up before or after the file name
        }
        :param data: A list of lists to put in a csv
        :param params: A dictionary containing all needed information
        """

        self.data = data
        self.task = task
        self.header = self.task.input_data_header
        self.file_path = self.task.file_storage
        self.file_name = self.task.file_name

        self.successful_run = False

        # Remove the header from the data set
        # if it is included in the data set
        if self.header is None:
            self.header = data[0]
            del self.data[0]

    def create_csv(self):
        """
        Write data to csv using pandas
        """
        try:
            # Convert List of Lists to DataFrame and write it to a CSV
            pd.DataFrame(self.data, columns=self.header) \
                .to_csv(os.path.join(self.file_path, self.file_name), index=False)
            self.successful_run = True
        except:
            # TODO create Exception Handling
            raise


# %% Excel Generator
class ExcelGenerator:
    """
    Generate Excel file with Data submitted
    """

    def __init__(self, data, file_name, sheet_name, range, file_path=None, dynamic_name=None):
        self.data = data
        if self.data:
            self.data_len = len(self.data)
            self.data_wid = max(len(x) for x in self.data)

        self.wb = None
        self.ws = None

        self.file_name = file_name
        self.sheet_name = sheet_name
        self.range_name = range

        self.file_path = file_path

    def create_workbook(self):
        """
        Creates workbook and writes data to the workbook
        """
        try:
            if '.xlsm' in self.file_name or '.xltm' in self.file_name:
                self.wb = load_workbook(self.file_path, keep_vba=True)
            else:
                if '.xlsx' not in self.file_name:
                    self.file_name = self.file_name + '.xlsx'
                self.wb = load_workbook(os.path.join(self.file_path, self.file_name))
        except Exception as e:
            self.wb = Workbook()

        sheet_names = self.wb.sheetnames
        if self.sheet_name in sheet_names:
            self.ws = self.wb[self.sheet_name]
        else:
            self.ws = self.wb.create_sheet(title=self.sheet_name)

        if self.data_len < self.ws.max_row:
            self.clear_sheet()
            self.wb.save(os.path.join(self.file_path, self.file_name))

        self.write_to_sheet()

        self.wb.save(os.path.join(self.file_path, self.file_name))

    def clear_sheet(self):
        for row in self.ws[self.range_name]:
            for cell in row:
                cell.value = None

    def write_to_sheet(self):
        """
        Writes a list to the worksheets named
        """
        for i, row in enumerate(self.ws[self.range_name]):
            if isinstance(self.data[i], tuple):
                self.data[i] = list(self.data[i])
            for j, cell in enumerate(row):
                cell.value = self.data[i][j]
